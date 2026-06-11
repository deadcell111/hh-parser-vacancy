from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.analysis import AnalysisState


class SaaSReportExporter:
    def export(self, state: AnalysisState, output_dir: str | Path) -> list[Path]:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        excel_path = output_dir / "HH_Analysis_Report.xlsx"
        pdf_path = output_dir / "HH_Analysis_Report.pdf"
        self.export_excel(state, excel_path)
        self.export_pdf(state, pdf_path)
        return [excel_path, pdf_path]

    def export_excel(self, state: AnalysisState, path: Path) -> None:
        frames = self._frames(state)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet, frame in frames.items():
                frame.to_excel(writer, sheet_name=sheet, index=False)
        self._format(path)

    def export_pdf(self, state: AnalysisState, path: Path) -> None:
        try:
            from PyQt6.QtCore import QRect
            from PyQt6.QtGui import QFont, QPainter, QPdfWriter
        except ImportError:
            path.with_suffix(".txt").write_text("PyQt6 is required for PDF export.", encoding="utf-8")
            return

        writer = QPdfWriter(str(path))
        painter = QPainter(writer)
        painter.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        painter.drawText(QRect(40, 40, 760, 60), 0, "HH Career Intelligence Report")
        painter.setFont(QFont("Arial", 10))
        lines = [
            f"Vacancies found: {state.summary.vacancies_found}",
            f"Average salary: {state.summary.average_salary}",
            f"Resume match: {state.summary.resume_match}%",
            f"Unique skills: {state.summary.unique_skills}",
        ]
        y = 120
        for line in lines:
            painter.drawText(40, y, line)
            y += 26
        painter.end()

    def _frames(self, state: AnalysisState) -> dict[str, pd.DataFrame]:
        vacancies = [
            {
                "Title": v.title,
                "Company": v.company,
                "City": v.city,
                "Salary": v.salary,
                "URL": v.url,
            }
            for v in state.vacancies
        ]
        skills = [
            {"Vacancy": v.display_name, "Skill": s["name"], "Category": s["category"]}
            for v in state.vacancies
            for s in v.skills
        ]
        market = [
            {"Metric": "Vacancies found", "Value": state.summary.vacancies_found},
            {"Metric": "Average salary", "Value": state.summary.average_salary},
            {"Metric": "Resume match", "Value": state.summary.resume_match},
            {"Metric": "Unique skills", "Value": state.summary.unique_skills},
        ]
        return {
            "Summary": pd.DataFrame(market),
            "Vacancies": pd.DataFrame(vacancies),
            "Skills": pd.DataFrame(skills),
            "Market Analysis": pd.DataFrame(state.summary.top_skills),
            "Resume Analysis": pd.DataFrame(self._dict_rows(state.resume_result)),
            "AI Advisor": pd.DataFrame(self._dict_rows(state.ai_advisor)),
            "Learning Roadmap": pd.DataFrame(self._roadmap_rows(state.roadmap)),
        }

    def _dict_rows(self, value: dict | None) -> list[dict[str, object]]:
        if not value:
            return []
        return [{"Section": key, "Value": item} for key, item in value.items()]

    def _roadmap_rows(self, value: dict | None) -> list[dict[str, object]]:
        if not value:
            return []
        weeks = value.get("weeks", []) if isinstance(value, dict) else []
        return weeks if isinstance(weeks, list) else []

    def _format(self, path: Path) -> None:
        workbook = load_workbook(path)
        fill = PatternFill("solid", fgColor="7C3AED")
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
            for column in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 14), 80)
        workbook.save(path)
