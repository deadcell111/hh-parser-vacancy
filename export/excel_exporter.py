from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from parser.models import VacancyData


class ExcelExporter:
    def export_all(
        self,
        vacancies: list[VacancyData],
        market_statistics: dict[str, pd.DataFrame],
        resume_result: dict[str, object] | None,
        output_dir: str | Path,
        excluded_vacancies: list[VacancyData] | None = None,
    ) -> list[Path]:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        frames = self.build_frames(vacancies, market_statistics, resume_result, excluded_vacancies or [])

        market_path = output_dir / "Market_Report.xlsx"
        with pd.ExcelWriter(market_path, engine="openpyxl") as writer:
            frames["vacancies"].to_excel(writer, sheet_name="Вакансии", index=False)
            frames["skills"].to_excel(writer, sheet_name="Навыки", index=False)
            frames["duties"].to_excel(writer, sheet_name="Обязанности", index=False)
            frames["requirements"].to_excel(writer, sheet_name="Требования", index=False)
            frames["top_skills"].to_excel(writer, sheet_name="ТОП навыков", index=False)
            frames["top_duties"].to_excel(writer, sheet_name="ТОП обязанностей", index=False)
            frames["top_requirements"].to_excel(writer, sheet_name="ТОП требований", index=False)
            frames["filtered"].to_excel(writer, sheet_name="Отфильтрованные вакансии", index=False)
        self._format_workbook(market_path)

        resume_path = output_dir / "Resume_Gap_Analysis.xlsx"
        frames["resume_gap_analysis"].to_excel(resume_path, index=False)
        self._format_workbook(resume_path)

        raw_path = output_dir / "raw_data.json"
        raw_path.write_text(
            json.dumps({key: frame.to_dict(orient="records") for key, frame in frames.items()}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return [market_path, resume_path, raw_path]

    def export_vacancies(self, vacancies: list[VacancyData], output_path: str | Path = "Market_Report.xlsx") -> Path:
        output_path = Path(output_path)
        frames = self.build_frames(vacancies, {}, None, [])
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            frames["vacancies"].to_excel(writer, sheet_name="Вакансии", index=False)
            frames["skills"].to_excel(writer, sheet_name="Навыки", index=False)
            frames["duties"].to_excel(writer, sheet_name="Обязанности", index=False)
            frames["requirements"].to_excel(writer, sheet_name="Требования", index=False)
        self._format_workbook(output_path)
        return output_path

    def export_skill_stats(self, stats: pd.DataFrame, output_path: str | Path = "top_skills.xlsx") -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        stats.to_excel(output_path, index=False)
        self._format_workbook(output_path)
        return output_path

    def build_frames(
        self,
        vacancies: list[VacancyData],
        market_statistics: dict[str, pd.DataFrame],
        resume_result: dict[str, object] | None,
        excluded_vacancies: list[VacancyData] | None = None,
    ) -> dict[str, pd.DataFrame]:
        excluded_vacancies = excluded_vacancies or []
        vacancy_rows = [
            {
                "ID": vacancy.vacancy_id,
                "Название вакансии": vacancy.title,
                "Компания": vacancy.company,
                "Зарплата": vacancy.salary,
                "Город": vacancy.city,
                "Релевантность": vacancy.relevance_score,
                "Ссылка": vacancy.url,
            }
            for vacancy in vacancies
        ]
        duty_rows = [
            {"Вакансия": vacancy.display_name, "Обязанность": item, "Ссылка": vacancy.url}
            for vacancy in vacancies
            for item in vacancy.responsibilities
        ]
        requirement_rows = [
            {"Вакансия": vacancy.display_name, "Требование": item, "Ссылка": vacancy.url}
            for vacancy in vacancies
            for item in vacancy.requirements
        ]
        skill_rows = [
            {"Вакансия": vacancy.display_name, "Навык": skill["name"], "Категория": skill["category"], "Ссылка": vacancy.url}
            for vacancy in vacancies
            for skill in vacancy.skills
        ]
        filtered_rows = [
            {
                "Вакансия": vacancy.display_name,
                "Релевантность": vacancy.relevance_score,
                "Причина": vacancy.exclusion_reason,
                "Детали": "; ".join(vacancy.relevance_details),
                "Ссылка": vacancy.url,
            }
            for vacancy in excluded_vacancies
        ]
        resume_rows = self._resume_rows(resume_result)

        return {
            "vacancies": pd.DataFrame(vacancy_rows),
            "skills": pd.DataFrame(skill_rows),
            "duties": pd.DataFrame(duty_rows),
            "requirements": pd.DataFrame(requirement_rows),
            "top_skills": market_statistics.get("skills", pd.DataFrame()),
            "top_duties": market_statistics.get("duties", pd.DataFrame()),
            "top_requirements": market_statistics.get("requirements", pd.DataFrame()),
            "filtered": pd.DataFrame(filtered_rows),
            "resume_gap_analysis": pd.DataFrame(resume_rows),
        }

    def _resume_rows(self, resume_result: dict[str, object] | None) -> list[dict[str, str]]:
        rows = []
        if not resume_result:
            return rows
        for skill in resume_result.get("matched_skills", []):
            rows.append({"Тип": "Найдено", "Навык": skill})
        for skill in resume_result.get("missing_skills", []):
            rows.append({"Тип": "Отсутствует", "Навык": skill})
        for recommendation in resume_result.get("recommendations", []):
            rows.append({"Тип": "Рекомендация", "Навык": recommendation})
        return rows

    def _format_workbook(self, path: Path) -> None:
        workbook = load_workbook(path)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for worksheet in workbook.worksheets:
            if worksheet.max_row == 0:
                continue
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 14), 80)
        workbook.save(path)
